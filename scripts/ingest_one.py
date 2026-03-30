#!/usr/bin/env python3
"""
AceCamp source-level downstream processor.

ROLE IN THE SKILL:
- This is NOT the top-level user entrypoint.
- It is the downstream processor that takes an already-built `source.md`
  and completes render / pdf / manifest-index sync / consistency fix / validation.

CALLERS:
- `ingest_from_open_page.py`
- future repair / re-render / re-index flows that already have source files
"""
import argparse
import re
import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from lib.config import load_config
from lib.error_log import append_error_log
from lib.source_rules import validate_source_file
from lib.html_renderer import render_file
from lib.index_updater import upsert_all
from lib.output_validator import run_check
from lib.pdf_renderer import render_rendered_to_pdf


def extract_source_metadata(source_path: Path):
    txt = source_path.read_text(encoding='utf-8')
    m_ind = re.search(r'^- 行业：(.+)$', txt, flags=re.M)
    industry = m_ind.group(1).strip() if m_ind else ''

    tags = []
    if '## 五、标签（页面可见）' in txt:
        sec = txt.split('## 五、标签（页面可见）', 1)[1]
        sec = sec.split('## 六、专家与作者信息', 1)[0]
        tags = [ln[2:].strip() for ln in sec.splitlines() if ln.startswith('- ')]

    author = ''
    co_publisher = ''
    if '## 六、专家与作者信息（页面可见）' in txt:
        sec = txt.split('## 六、专家与作者信息（页面可见）', 1)[1]
        sec = sec.split('## 七、补充说明', 1)[0]
        m_author = re.search(r'^- 发布人：(.+)$', sec, flags=re.M)
        if m_author:
            author = m_author.group(1).strip()
        m_co = re.search(r'^- 联合发布人：(.+)$', sec, flags=re.M)
        if m_co:
            co_publisher = m_co.group(1).strip()

    return industry, tags, author, co_publisher


def auto_fix_consistency(article_id: str, source_path: Path, strict_consistency: bool, source_url: str):
    ws_root = Path(__file__).resolve().parents[3]
    manifest_path = ws_root / 'acecamp-raw' / 'index' / 'manifest.jsonl'
    index_path = ws_root / 'acecamp-raw' / 'index' / '索引.xlsx'

    src_industry, src_tags, _src_author, _src_co = extract_source_metadata(source_path)

    # manifest patch
    rows = []
    mismatches = []
    for ln in manifest_path.read_text(encoding='utf-8').splitlines():
        if not ln.strip():
            continue
        r = json.loads(ln)
        if str(r.get('article_id', '')) == str(article_id):
            if r.get('industry', '') != src_industry:
                mismatches.append(('industry', r.get('industry', ''), src_industry))
                r['industry'] = src_industry
            if r.get('tags', []) != src_tags:
                mismatches.append(('tags', r.get('tags', []), src_tags))
                r['tags'] = src_tags
        rows.append(r)
    if mismatches:
        manifest_path.write_text(''.join(json.dumps(r, ensure_ascii=False) + '\n' for r in rows), encoding='utf-8')

    # index patch
    wb = load_workbook(index_path)
    ws = wb.active
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for rr in range(2, ws.max_row + 1):
        if str(ws.cell(rr, headers['article_id']).value or '') == str(article_id):
            ws.cell(rr, headers['行业']).value = src_industry
            ws.cell(rr, headers['标签']).value = ', '.join(src_tags)
            break
    wb.save(index_path)

    if mismatches:
        print(f"WARNING DataMismatch auto-fixed for article {article_id}: {mismatches}")
        append_error_log(ws_root, article_id, 'post_ingest_check', 'DataMismatch', 'industry/tags mismatch auto-fixed', source_url, {
            'action_taken': 'auto_fixed',
            'mismatches': mismatches,
        })
        if strict_consistency:
            raise RuntimeError('DataMismatch detected and strict_consistency=true')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--source-path', required=True)
    ap.add_argument('--rendered-path', required=True)
    ap.add_argument('--source-url', required=True)
    ap.add_argument('--record-time', required=True)
    ap.add_argument('--article-id', required=True)
    ap.add_argument('--article-date', required=True)
    ap.add_argument('--provider', default='acecamptech')
    ap.add_argument('--article-title', required=True)
    ap.add_argument('--crawl-date', required=True)
    ap.add_argument('--content-type', default='')
    ap.add_argument('--industry', default='')
    ap.add_argument('--tags', default='')
    ap.add_argument('--author', default='')
    ap.add_argument('--co-publisher', default='')
    ap.add_argument('--allow-empty-body', action='store_true')
    ap.add_argument('--min-body-chars', type=int, default=None)
    ap.add_argument('--strict-consistency', action='store_true')
    args = ap.parse_args()

    cfg = load_config()
    min_body_chars = args.min_body_chars if args.min_body_chars is not None else int(cfg.get('min_body_chars', 300))
    strict_consistency = args.strict_consistency or bool(cfg.get('strict_consistency', False))

    source_path = Path(args.source_path)
    rendered_path = Path(args.rendered_path)

    validate_source_file(source_path, allow_empty_body=args.allow_empty_body, min_body_chars=min_body_chars)

    render_file(
        input_md=str(source_path),
        output_html=str(rendered_path),
        source_url=args.source_url,
        record_time=args.record_time,
        article_id=args.article_id,
    )

    # 一体化验收：source 与 rendered 格式保真校验
    run_check(
        source_path=source_path,
        rendered_path=rendered_path,
        allow_empty_body=args.allow_empty_body,
        min_body_chars=min_body_chars,
    )

    # 直接生成 PDF 到 acecamp-raw/share 目录，避免 rendered 目录同时堆放 html+pdf
    share_dir = rendered_path.parents[1] / 'share'
    share_dir.mkdir(parents=True, exist_ok=True)
    share_pdf_path = share_dir / rendered_path.with_suffix('.pdf').name
    render_rendered_to_pdf(
        rendered_html=rendered_path,
        rendered_pdf=share_pdf_path,
    )

    src_industry, src_tags, src_author, src_co_publisher = extract_source_metadata(source_path)
    tags = src_tags if src_tags else [t.strip() for t in args.tags.replace('，', ',').split(',') if t.strip()]
    industry = src_industry if src_industry else args.industry
    author = src_author if src_author else args.author
    co_publisher = src_co_publisher if src_co_publisher else args.co_publisher

    entry = {
        'article_date': args.article_date,
        'provider': args.provider,
        'article_id': str(args.article_id),
        'article_title': args.article_title,
        'crawl_date': args.crawl_date,
        'source_url': args.source_url,
        'record_time': args.record_time,
        'source_path': str(source_path),
        'rendered_path': str(rendered_path),
        'status': 'ok',
        'content_type': args.content_type,
        'industry': industry,
        'tags': tags,
        'author': author,
        'co_publisher': co_publisher,
    }
    upsert_all(entry)

    auto_fix_consistency(args.article_id, source_path, strict_consistency, args.source_url)

    print('INGEST_ONE_OK', args.article_id)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        # best-effort parse for context
        article_id = ''
        source_url = ''
        argv = ' '.join(sys.argv)
        m = re.search(r'--article-id\s+(\S+)', argv)
        if m:
            article_id = m.group(1)
        u = re.search(r'--source-url\s+(\S+)', argv)
        if u:
            source_url = u.group(1)
        ws_root = Path(__file__).resolve().parents[3]
        append_error_log(ws_root, article_id, 'ingest_one', type(e).__name__, str(e), source_url)
        print(f'INGEST_ONE_FAIL: {e}')
        sys.exit(1)
