import json
from pathlib import Path
from openpyxl import load_workbook, Workbook


def upsert_manifest(manifest_path: Path, entry: dict):
    rows = []
    if manifest_path.exists():
        for ln in manifest_path.read_text(encoding='utf-8').splitlines():
            if ln.strip():
                rows.append(json.loads(ln))

    updated = False
    for i, r in enumerate(rows):
        if str(r.get('article_id', '')) == str(entry['article_id']):
            rows[i] = entry
            updated = True
            break
    if not updated:
        rows.append(entry)

    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(''.join(json.dumps(r, ensure_ascii=False) + '\n' for r in rows), encoding='utf-8')


def upsert_excel(index_path: Path, entry: dict):
    index_path.parent.mkdir(parents=True, exist_ok=True)
    if index_path.exists():
        wb = load_workbook(index_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['article_date', 'provider', 'article_id', 'article_title', 'crawl_date', 'source_url', 'record_time', 'source_path', 'rendered_path', 'status', '内容类型', '行业', '标签', '发布人', '联合发布人'])

    required_headers = ['article_date', 'provider', 'article_id', 'article_title', 'crawl_date', 'source_url', 'record_time', 'source_path', 'rendered_path', 'status', '内容类型', '行业', '标签', '发布人', '联合发布人']
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for h in required_headers:
        if h not in headers:
            col = ws.max_column + 1
            ws.cell(1, col).value = h
            headers[h] = col
    row_idx = None
    for rr in range(2, ws.max_row + 1):
        if str(ws.cell(rr, headers['article_id']).value or '') == str(entry['article_id']):
            row_idx = rr
            break
    if row_idx is None:
        row_idx = ws.max_row + 1

    values = {
        'article_date': entry.get('article_date', ''),
        'provider': entry.get('provider', ''),
        'article_id': entry.get('article_id', ''),
        'article_title': entry.get('article_title', ''),
        'crawl_date': entry.get('crawl_date', ''),
        'source_url': entry.get('source_url', ''),
        'record_time': entry.get('record_time', ''),
        'source_path': entry.get('source_path', ''),
        'rendered_path': entry.get('rendered_path', ''),
        'status': entry.get('status', ''),
        '内容类型': entry.get('content_type', ''),
        '行业': entry.get('industry', ''),
        '标签': ', '.join(entry.get('tags', [])),
        '发布人': entry.get('author', ''),
        '联合发布人': entry.get('co_publisher', ''),
    }
    for k, v in values.items():
        ws.cell(row_idx, headers[k]).value = v

    wb.save(index_path)


def upsert_all(entry: dict, ws_root: Path = None):
    if ws_root is None:
        ws_root = Path(__file__).resolve().parents[4]
    manifest_path = ws_root / 'acecamp-raw' / 'index' / 'manifest.jsonl'
    index_path = ws_root / 'acecamp-raw' / 'index' / '索引.xlsx'
    upsert_manifest(manifest_path, entry)
    upsert_excel(index_path, entry)
